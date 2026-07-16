import os
import re
import tempfile
from flask import Flask, render_template_string, request, send_file
import pdfplumber
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

app = Flask(__name__)
# 限制上传文件最大为 50MB，可自行调整
app.config['MAX_CONTENT_LENGTH'] = 50 * 1024 * 1024

def custom_extract_table(page):
    tables = page.find_tables()
    if not tables:
        return []
    all_tables = []
    for table in tables:
        grid = table.extract()
        if grid:
            cleaned = [[cell.strip() if cell else '' for cell in row] for row in grid]
            all_tables.append(cleaned)
    return all_tables

def merge_split_rows(table):
    if not table:
        return table
    merged = []
    for row in table:
        if not row:
            continue
        filled = [i for i, cell in enumerate(row) if cell and cell.strip()]
        if len(filled) == 1 and merged:
            col_idx = filled[0]
            prev = merged[-1]
            while len(prev) <= col_idx:
                prev.append('')
            prev[col_idx] = (prev[col_idx] or '') + '\n' + row[col_idx].strip()
        elif len(filled) == 0:
            continue
        else:
            merged.append(row)
    return merged

def fallback_extract_table(page, x_tol=8, y_tol=9):
    words = page.extract_words(keep_blank_chars=False)
    if not words:
        return []
    words = sorted(words, key=lambda w: (round(w['top'], 1), w['x0']))
    rows = []
    current_row = []
    last_top = None
    for w in words:
        if last_top is None or abs(w['top'] - last_top) > y_tol:
            if current_row:
                rows.append(current_row)
            current_row = [w]
            last_top = w['top']
        else:
            current_row.append(w)
    if current_row:
        rows.append(current_row)
    table = []
    for row in rows:
        row.sort(key=lambda w: w['x0'])
        cols = []
        cur_text = row[0]['text']
        last_x1 = row[0]['x1']
        for w in row[1:]:
            if w['x0'] - last_x1 > x_tol:
                cols.append(cur_text.strip())
                cur_text = w['text']
            else:
                cur_text += ' ' + w['text']
            last_x1 = w['x1']
        cols.append(cur_text.strip())
        table.append(cols)
    max_cols = max(len(r) for r in table) if table else 0
    for r in table:
        while len(r) < max_cols:
            r.append('')
    return [table]

def beautify_excel_sheet(ws):
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    # 所有单元格统一居中
    center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            cell.border = thin_border
            cell.alignment = center_alignment
            if cell.row == 1:
                cell.font = header_font
                cell.fill = header_fill
    
    # 自动列宽
    for col_cells in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value:
                for line in str(cell.value).split('\n'):
                    length = sum(2 if ord(c) > 127 else 1 for c in line)
                    if length > max_length:
                        max_length = length
        adjusted = min(max_length + 2, 60)
        ws.column_dimensions[col_letter].width = max(adjusted, 8)
    ws.freeze_panes = 'A2'

def process_pdf_to_excel(pdf_path, start_page, end_page, output_path):
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        has_data = False
        with pdfplumber.open(pdf_path) as pdf:
            total = len(pdf.pages)
            start = max(1, start_page)
            end = min(end_page, total)
            pages = list(range(start, end + 1))
            
            for num in pages:
                page = pdf.pages[num - 1]
                tables = custom_extract_table(page)
                if not tables:
                    tables = fallback_extract_table(page)
                    tables = [merge_split_rows(t) for t in tables]
                for idx, table in enumerate(tables):
                    if not table:
                        continue
                    cleaned = [[(c or '').strip() for c in row] for row in table]
                    cleaned = [r for r in cleaned if any(c for c in r)]
                    if not cleaned:
                        continue
                    # 处理序号列
                    if len(cleaned) > 1:
                        first_col = [r[0] for r in cleaned]
                        if sum(1 for c in first_col if re.match(r'^\d+\.\s*$', c)) > len(first_col) * 0.5:
                            for row in cleaned:
                                m = re.match(r'^(\d+)\.\s*$', row[0])
                                if m:
                                    row[0] = m.group(1)
                    df = pd.DataFrame(cleaned[1:], columns=cleaned[0]) if len(cleaned) > 1 else pd.DataFrame(cleaned)
                    sheet = f'Page{num}_T{idx+1}'
                    df.to_excel(writer, sheet_name=sheet, index=False)
                    has_data = True
        
        if not has_data:
            pd.DataFrame({'提示': ['未提取到表格']}).to_excel(writer, sheet_name='无数据', index=False)
    
    # 美化
    wb = load_workbook(output_path)
    for ws in wb.worksheets:
        beautify_excel_sheet(ws)
    wb.save(output_path)

# ---------- 前端 HTML（含拖拽上传区域） ----------
HTML_TEMPLATE = '''
<!DOCTYPE html>
<html lang="zh">
<head>
<meta charset="UTF-8">
<title>PDF 拖拽转 Excel</title>
<style>
    body { font-family: "Microsoft YaHei", sans-serif; padding: 30px; }
    h2 { color: #333; }
    #drop-area {
        border: 2px dashed #aaa;
        border-radius: 10px;
        padding: 40px;
        text-align: center;
        background: #f9f9f9;
        margin-bottom: 20px;
        transition: background 0.3s;
    }
    #drop-area.highlight { background: #e1f5fe; border-color: #03a9f4; }
    input[type="file"] { display: none; }
    button {
        padding: 8px 20px; font-size: 16px; margin-top: 10px;
        cursor: pointer; background: #1976d2; color: white;
        border: none; border-radius: 5px;
    }
    button:hover { background: #1565c0; }
    .row { margin: 10px 0; }
    input[type="number"] { width: 80px; padding: 5px; }
    .message { margin: 15px 0; color: #d32f2f; }
    .success { color: #388e3c; }
    a { color: #1976d2; }
</style>
</head>
<body>
<h2>📄 PDF 拖拽转 Excel</h2>

<div id="drop-area">
    <p>把 PDF 文件拖到这里<br>或 <a href="#" onclick="document.getElementById('fileInput').click()">点击选择文件</a></p>
    <input type="file" id="fileInput" accept=".pdf" />
</div>

<div id="fileInfo" style="margin-bottom:15px; font-weight:bold;"></div>

<div class="row">
    <label>起始页：<input type="number" id="startPage" min="1" value="1"></label>
    <label>结束页：<input type="number" id="endPage" min="1" value="1"></label>
    <span style="font-size:0.9em; color:#888;">（留空或 0 表示全部）</span>
</div>

<button onclick="uploadPDF()">开始转换</button>
<div id="message" class="message"></div>
<div id="downloadLink"></div>

<script>
    const dropArea = document.getElementById('drop-area');
    const fileInput = document.getElementById('fileInput');
    let selectedFile = null;

    // 阻止浏览器默认拖拽行为
    ['dragenter', 'dragover', 'dragleave', 'drop'].forEach(eventName => {
        dropArea.addEventListener(eventName, preventDefaults, false);
        document.body.addEventListener(eventName, preventDefaults, false);
    });
    function preventDefaults(e) { e.preventDefault(); e.stopPropagation(); }

    // 高亮拖拽区域
    ['dragenter', 'dragover'].forEach(eventName => {
        dropArea.addEventListener(eventName, () => dropArea.classList.add('highlight'), false);
    });
    ['dragleave', 'drop'].forEach(eventName => {
        dropArea.addEventListener(eventName, () => dropArea.classList.remove('highlight'), false);
    });

    // 处理拖拽文件
    dropArea.addEventListener('drop', handleDrop, false);
    function handleDrop(e) {
        const dt = e.dataTransfer;
        const files = dt.files;
        if (files.length > 0) {
            selectedFile = files[0];
            fileInput.files = files; // 同步给 input
            showFileInfo();
        }
    }

    // 处理点击选择文件
    fileInput.addEventListener('change', function(e) {
        if (this.files.length > 0) {
            selectedFile = this.files[0];
            showFileInfo();
        }
    });

    function showFileInfo() {
        if (selectedFile) {
            document.getElementById('fileInfo').innerHTML = `✅ 已选择：${selectedFile.name} (${(selectedFile.size/1024/1024).toFixed(2)} MB)`;
        }
    }

    async function uploadPDF() {
        if (!selectedFile) {
            alert('请先拖拽或选择一个 PDF 文件！');
            return;
        }
        const startPage = document.getElementById('startPage').value || 0;
        const endPage = document.getElementById('endPage').value || 0;
        const formData = new FormData();
        formData.append('pdf', selectedFile);
        formData.append('start', startPage);
        formData.append('end', endPage);

        document.getElementById('message').innerHTML = '⏳ 正在转换中，请耐心等待几分钟...';
        document.getElementById('downloadLink').innerHTML = '';

        try {
            const response = await fetch('/convert', {
                method: 'POST',
                body: formData
            });
            if (!response.ok) {
                const error = await response.json();
                throw new Error(error.message || '转换失败');
            }
            const blob = await response.blob();
            const url = window.URL.createObjectURL(blob);
            const a = document.createElement('a');
            a.href = url;
            a.download = '表格提取结果.xlsx';
            document.getElementById('downloadLink').innerHTML = `<p class="success">✅ 转换成功！<a href="${url}" download="表格提取结果.xlsx">点击下载 Excel 文件</a></p>`;
            document.getElementById('message').innerHTML = '';
        } catch (err) {
            document.getElementById('message').innerHTML = '❌ 错误：' + err.message;
        }
    }
</script>
</body>
</html>
'''

# ---------- 路由 ----------
@app.route('/')
def index():
    return render_template_string(HTML_TEMPLATE)

@app.route('/convert', methods=['POST'])
def convert():
    if 'pdf' not in request.files:
        return {'message': '没有上传文件'}, 400
    file = request.files['pdf']
    if file.filename == '':
        return {'message': '文件名为空'}, 400
    if not file.filename.lower().endswith('.pdf'):
        return {'message': '只允许上传 PDF 文件'}, 400

    start_page = request.form.get('start', 0, type=int)
    end_page = request.form.get('end', 0, type=int)

    # 保存上传的文件到临时位置
    temp_dir = tempfile.mkdtemp()
    pdf_path = os.path.join(temp_dir, file.filename)
    file.save(pdf_path)

    # 输出 Excel 临时文件
    out_path = os.path.join(temp_dir, 'output.xlsx')

    try:
        # 若起始页未指定或为0，则提取全部
        with pdfplumber.open(pdf_path) as pdf:
            total_pages = len(pdf.pages)
        if start_page <= 0:
            start_page = 1
        if end_page <= 0 or end_page > total_pages:
            end_page = total_pages

        process_pdf_to_excel(pdf_path, start_page, end_page, out_path)
        # 返回 Excel 文件并清理临时目录
        response = send_file(out_path, as_attachment=True, download_name='表格提取结果.xlsx')
        # 注册清理回调（简单做法：在发送后删除临时目录，此处使用 after_request）
        @response.call_on_close
        def cleanup():
            import shutil
            shutil.rmtree(temp_dir, ignore_errors=True)
        return response
    except Exception as e:
        import shutil
        shutil.rmtree(temp_dir, ignore_errors=True)
        return {'message': f'转换失败：{str(e)}'}, 500

if __name__ == '__main__':
    app.run()