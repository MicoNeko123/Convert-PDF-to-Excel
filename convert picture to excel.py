import os
import re
import tempfile
from flask import Flask, render_template_string, request, send_file
import pdfplumber
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
import easyocr
from PIL import Image
import numpy as np

app = Flask(__name__)
app.config['MAX_CONTENT_LENGTH'] = 50 * 1024 * 1024  # 最大上传 50MB

# ==================== PDF 表格提取函数（保持不变） ====================
def custom_extract_table(page):
    tables = page.find_tables()
    if not tables:
        return []
    all_tables = []
    for table in tables:
        grid = table.extract()
        if grid:
            cleaned = [[cell.strip() if cell else '' for cell in row] for row in grid]
            all_tables.append(cleaned)
    return all_tables

def merge_split_rows(table):
    if not table:
        return table
    merged = []
    for row in table:
        if not row:
            continue
        filled = [i for i, cell in enumerate(row) if cell and cell.strip()]
        if len(filled) == 1 and merged:
            col_idx = filled[0]
            prev = merged[-1]
            while len(prev) <= col_idx:
                prev.append('')
            prev[col_idx] = (prev[col_idx] or '') + '\n' + row[col_idx].strip()
        elif len(filled) == 0:
            continue
        else:
            merged.append(row)
    return merged

def fallback_extract_table(page, x_tol=8, y_tol=9):
    words = page.extract_words(keep_blank_chars=False)
    if not words:
        return []
    words = sorted(words, key=lambda w: (round(w['top'], 1), w['x0']))
    rows = []
    current_row = []
    last_top = None
    for w in words:
        if last_top is None or abs(w['top'] - last_top) > y_tol:
            if current_row:
                rows.append(current_row)
            current_row = [w]
            last_top = w['top']
        else:
            current_row.append(w)
    if current_row:
        rows.append(current_row)
    table = []
    for row in rows:
        row.sort(key=lambda w: w['x0'])
        cols = []
        cur_text = row[0]['text']
        last_x1 = row[0]['x1']
        for w in row[1:]:
            if w['x0'] - last_x1 > x_tol:
                cols.append(cur_text.strip())
                cur_text = w['text']
            else:
                cur_text += ' ' + w['text']
            last_x1 = w['x1']
        cols.append(cur_text.strip())
        table.append(cols)
    max_cols = max(len(r) for r in table) if table else 0
    for r in table:
        while len(r) < max_cols:
            r.append('')
    return [table]

# ==================== Excel 美化（已居中） ====================
def beautify_excel_sheet(ws):
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            cell.border = thin_border
            cell.alignment = center_alignment
            if cell.row == 1:
                cell.font = header_font
                cell.fill = header_fill
    
    for col_cells in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value:
                for line in str(cell.value).split('\n'):
                    length = sum(2 if ord(c) > 127 else 1 for c in line)
                    if length > max_length:
                        max_length = length
        adjusted = min(max_length + 2, 60)
        ws.column_dimensions[col_letter].width = max(adjusted, 8)
    ws.freeze_panes = 'A2'

# ==================== PDF 转 Excel 流程 ====================
def process_pdf_to_excel(pdf_path, start_page, end_page, output_path):
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        has_data = False
        with pdfplumber.open(pdf_path) as pdf:
            total = len(pdf.pages)
            start = max(1, start_page)
            end = min(end_page, total)
            pages = list(range(start, end + 1))
            
            for num in pages:
                page = pdf.pages[num - 1]
                tables = custom_extract_table(page)
                if not tables:
                    tables = fallback_extract_table(page)
                    tables = [merge_split_rows(t) for t in tables]
                for idx, table in enumerate(tables):
                    if not table:
                        continue
                    cleaned = [[(c or '').strip() for c in row] for row in table]
                    cleaned = [r for r in cleaned if any(c for c in r)]
                    if not cleaned:
                        continue
                    if len(cleaned) > 1:
                        first_col = [r[0] for r in cleaned]
                        if sum(1 for c in first_col if re.match(r'^\d+\.\s*$', c)) > len(first_col) * 0.5:
                            for row in cleaned:
                                m = re.match(r'^(\d+)\.\s*$', row[0])
                                if m:
                                    row[0] = m.group(1)
                    df = pd.DataFrame(cleaned[1:], columns=cleaned[0]) if len(cleaned) > 1 else pd.DataFrame(cleaned)
                    sheet = f'Page{num}_T{idx+1}'
                    df.to_excel(writer, sheet_name=sheet, index=False)
                    has_data = True
        
        if not has_data:
            pd.DataFrame({'提示': ['未提取到表格']}).to_excel(writer, sheet_name='无数据', index=False)
    
    wb = load_workbook(output_path)
    for ws in wb.worksheets:
        beautify_excel_sheet(ws)
    wb.save(output_path)

# ==================== 图片转表格（EasyOCR） ====================
# 全局初始化一个 EasyOCR reader（懒加载，避免每次请求都创建）
reader = None

def get_reader():
    global reader
    if reader is None:
        # 首次运行会下载模型（约 300MB），请耐心等待
        reader = easyocr.Reader(['ch_sim', 'en'], gpu=False)
    return reader

def image_to_table(image_path):
    """使用 EasyOCR 识别图片中的文字，并按坐标聚类成表格"""
    reader = get_reader()
    results = reader.readtext(image_path)
    if not results:
        return []
    
    # results 格式: [([[x1,y1],[x2,y2],[x3,y3],[x4,y4]], text, confidence), ...]
    # 提取中心点 y 坐标，按 y 排序
    items = []
    for bbox, text, conf in results:
        # bbox 是四个点的列表，计算左上角和右下角
        x_coords = [p[0] for p in bbox]
        y_coords = [p[1] for p in bbox]
        x0, x1 = min(x_coords), max(x_coords)
        y0, y1 = min(y_coords), max(y_coords)
        center_y = (y0 + y1) / 2
        items.append({'x0': x0, 'x1': x1, 'y0': y0, 'y1': y1, 'center_y': center_y, 'text': text})
    
    if not items:
        return []
    
    # 按中心 y 排序
    items.sort(key=lambda it: it['center_y'])
    
    # 按行分组：当前项与上一项的 center_y 差 > 行高阈值（取平均行高的 0.6 倍）
    # 先估算平均行高
    heights = [it['y1'] - it['y0'] for it in items]
    avg_height = np.mean(heights) if heights else 10
    row_gap_threshold = avg_height * 0.6
    
    rows = []
    current_row = []
    last_center_y = None
    for it in items:
        if last_center_y is None or (it['center_y'] - last_center_y) > row_gap_threshold:
            if current_row:
                rows.append(current_row)
            current_row = [it]
            last_center_y = it['center_y']
        else:
            current_row.append(it)
            # 更新中心 y 为当前行的平均（可选）
            last_center_y = np.mean([r['center_y'] for r in current_row])
    if current_row:
        rows.append(current_row)
    
    # 每行内按 x0 排序，然后根据 x 间隙分列
    col_gap_threshold = avg_height * 1.2  # 列间隙阈值，可根据实际情况调整
    table = []
    for row in rows:
        row.sort(key=lambda it: it['x0'])
        cols = []
        cur_text = row[0]['text']
        last_x1 = row[0]['x1']
        for it in row[1:]:
            if it['x0'] - last_x1 > col_gap_threshold:
                cols.append(cur_text.strip())
                cur_text = it['text']
            else:
                cur_text += ' ' + it['text']
            last_x1 = it['x1']
        cols.append(cur_text.strip())
        table.append(cols)
    
    # 补齐列数
    max_cols = max(len(r) for r in table) if table else 0
    for r in table:
        while len(r) < max_cols:
            r.append('')
    
    return [table]  # 保持与 PDF 相同的返回格式（列表套表格）

def process_image_to_excel(image_path, output_path):
    """图片 -> Excel,包含美化和清洗"""
    tables = image_to_table(image_path)
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        has_data = False
        for idx, table in enumerate(tables):
            if not table:
                continue
            cleaned = [[(c or '').strip() for c in row] for row in table]
            cleaned = [r for r in cleaned if any(c for c in r)]
            if not cleaned:
                continue
            # 可选：处理序号列（与 PDF 保持一致）
            if len(cleaned) > 1:
                first_col = [r[0] for r in cleaned]
                if sum(1 for c in first_col if re.match(r'^\d+\.\s*$', c)) > len(first_col) * 0.5:
                    for row in cleaned:
                        m = re.match(r'^(\d+)\.\s*$', row[0])
                        if m:
                            row[0] = m.group(1)
            df = pd.DataFrame(cleaned[1:], columns=cleaned[0]) if len(cleaned) > 1 else pd.DataFrame(cleaned)
            sheet = f'Table_{idx+1}'
            df.to_excel(writer, sheet_name=sheet, index=False)
            has_data = True
        if not has_data:
            pd.DataFrame({'提示': ['未检测到表格']}).to_excel(writer, sheet_name='无数据', index=False)
    
    wb = load_workbook(output_path)
    for ws in wb.worksheets:
        beautify_excel_sheet(ws)
    wb.save(output_path)

# ==================== 前端 HTML（支持图片拖拽） ====================
HTML_TEMPLATE = '''
<!DOCTYPE html>
<html lang="zh">
<head>
<meta charset="UTF-8">
<title> 图片 表格提取</title>
<style>
    body { font-family: "Microsoft YaHei", sans-serif; padding: 30px; }
    h2 { color: #333; }
    #drop-area {
        border: 2px dashed #aaa;
        border-radius: 10px;
        padding: 40px;
        text-align: center;
        background: #f9f9f9;
        margin-bottom: 20px;
        transition: background 0.3s;
    }
    #drop-area.highlight { background: #e1f5fe; border-color: #03a9f4; }
    input[type="file"] { display: none; }
    button {
        padding: 8px 20px; font-size: 16px; margin-top: 10px;
        cursor: pointer; background: #1976d2; color: white;
        border: none; border-radius: 5px;
    }
    button:hover { background: #1565c0; }
    .row { margin: 10px 0; }
    input[type="number"] { width: 80px; padding: 5px; }
    .message { margin: 15px 0; color: #d32f2f; }
    .success { color: #388e3c; }
    a { color: #1976d2; }
    .note { font-size:0.9em; color:#888; }
</style>
</head>
<body>
<h2> 图片 转 Excel(拖拽上传)</h2>

<div id="drop-area">
    <p>把 图片文件拖到这里<br>或 <a href="#" onclick="document.getElementById('fileInput').click()">点击选择文件</a></p>
    <input type="file" id="fileInput" accept=".pdf,.png,.jpg,.jpeg,.bmp,.tiff" />
</div>

<div id="fileInfo" style="margin-bottom:15px; font-weight:bold;"></div>

<div class="row" id="pageRange" style="display:none;">
    <label>起始页：<input type="number" id="startPage" min="1" value="1"></label>
    <label>结束页：<input type="number" id="endPage" min="1" value="1"></label>
    <span class="note">（留空或 0 表示全部）</span>
</div>

<button onclick="uploadFile()">开始转换</button>
<div id="message" class="message"></div>
<div id="downloadLink"></div>
<div class="note">※ 图片转表格首次使用需下载模型(约300MB),请耐心等待。转换过程可能需要1~2分钟。</div>

<script>
    const dropArea = document.getElementById('drop-area');
    const fileInput = document.getElementById('fileInput');
    let selectedFile = null;

    ['dragenter', 'dragover', 'dragleave', 'drop'].forEach(eventName => {
        dropArea.addEventListener(eventName, preventDefaults, false);
        document.body.addEventListener(eventName, preventDefaults, false);
    });
    function preventDefaults(e) { e.preventDefault(); e.stopPropagation(); }

    ['dragenter', 'dragover'].forEach(eventName => {
        dropArea.addEventListener(eventName, () => dropArea.classList.add('highlight'), false);
    });
    ['dragleave', 'drop'].forEach(eventName => {
        dropArea.addEventListener(eventName, () => dropArea.classList.remove('highlight'), false);
    });

    dropArea.addEventListener('drop', handleDrop, false);
    function handleDrop(e) {
        const dt = e.dataTransfer;
        const files = dt.files;
        if (files.length > 0) {
            selectedFile = files[0];
            fileInput.files = files;
            showFileInfo();
        }
    }

    fileInput.addEventListener('change', function(e) {
        if (this.files.length > 0) {
            selectedFile = this.files[0];
            showFileInfo();
        }
    });

    function showFileInfo() {
        if (selectedFile) {
            const isPDF = selectedFile.name.toLowerCase().endsWith('.pdf');
            document.getElementById('fileInfo').innerHTML = `✅ 已选择：${selectedFile.name} (${(selectedFile.size/1024/1024).toFixed(2)} MB)`;
            document.getElementById('pageRange').style.display = isPDF ? 'block' : 'none';
        }
    }

    async function uploadFile() {
        if (!selectedFile) {
            alert('请先拖拽或选择一个文件！');
            return;
        }
        const formData = new FormData();
        formData.append('file', selectedFile);
        // 如果是 PDF,带上页码范围
        if (selectedFile.name.toLowerCase().endsWith('.pdf')) {
            formData.append('start', document.getElementById('startPage').value || 0);
            formData.append('end', document.getElementById('endPage').value || 0);
        }

        document.getElementById('message').innerHTML = '⏳ 正在转换中，请耐心等待（可能需要几分钟）...';
        document.getElementById('downloadLink').innerHTML = '';

        try {
            const response = await fetch('/convert', {
                method: 'POST',
                body: formData
            });
            if (!response.ok) {
                const error = await response.json();
                throw new Error(error.message || '转换失败');
            }
            const blob = await response.blob();
            const url = window.URL.createObjectURL(blob);
            const a = document.createElement('a');
            a.href = url;
            a.download = '表格提取结果.xlsx';
            document.getElementById('downloadLink').innerHTML = `<p class="success">✅ 转换成功！<a href="${url}" download="表格提取结果.xlsx">点击下载 Excel 文件</a></p>`;
            document.getElementById('message').innerHTML = '';
        } catch (err) {
            document.getElementById('message').innerHTML = '❌ 错误：' + err.message;
        }
    }
</script>
</body>
</html>
'''

# ==================== 路由 ====================
@app.route('/')
def index():
    return render_template_string(HTML_TEMPLATE)

@app.route('/convert', methods=['POST'])
def convert():
    if 'file' not in request.files:
        return {'message': '没有上传文件'}, 400
    file = request.files['file']
    if file.filename == '':
        return {'message': '文件名为空'}, 400

    filename = file.filename.lower()
    is_pdf = filename.endswith('.pdf')
    is_image = filename.endswith(('.png', '.jpg', '.jpeg', '.bmp', '.tiff'))
    if not (is_pdf or is_image):
        return {'message': '仅支持 PDF 或常见图片格式(png/jpg/bmp/tiff)'}, 400

    temp_dir = tempfile.mkdtemp()
    file_path = os.path.join(temp_dir, file.filename)
    file.save(file_path)
    out_path = os.path.join(temp_dir, 'output.xlsx')

    try:
        if is_pdf:
            start_page = request.form.get('start', 0, type=int)
            end_page = request.form.get('end', 0, type=int)
            with pdfplumber.open(file_path) as pdf:
                total = len(pdf.pages)
            if start_page <= 0: start_page = 1
            if end_page <= 0 or end_page > total: end_page = total
            process_pdf_to_excel(file_path, start_page, end_page, out_path)
        else:  # 图片
            process_image_to_excel(file_path, out_path)

        response = send_file(out_path, as_attachment=True, download_name='表格提取结果.xlsx')
        @response.call_on_close
        def cleanup():
            import shutil
            shutil.rmtree(temp_dir, ignore_errors=True)
        return response
    except Exception as e:
        import shutil
        shutil.rmtree(temp_dir, ignore_errors=True)
        return {'message': f'转换失败：{str(e)}'}, 500

if __name__ == '__main__':
    print("正在启动服务... 首次运行如卡顿，可能正在下载 EasyOCR 模型，请耐心等待。")
    app.run(host='0.0.0.0', port=5020, debug=True)